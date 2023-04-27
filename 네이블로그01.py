# 네이블로그01.py

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Create a new Excel workbook
wb = Workbook()

# Create a new worksheet in the workbook
ws = wb.active
ws.title = 'Blog Posts'

# Get the search keyword from the user
search_keyword = input('Enter a search keyword: ')

# Retrieve 100 pages of search results
for page in range(1, 101):
    # URL of the Naver search result page for the given keyword and page number
    url = f'https://search.naver.com/search.naver?where=view&sm=tab_jum&query={search_keyword}&start={page * 10 - 9}'

    # Send a GET request to the URL and get the HTML response
    response = requests.get(url)

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find all the blog titles and URLs using their HTML tags and classes
    blog_titles = soup.find_all('a', class_='api_txt_lines total_tit')
    blog_urls = soup.find_all('a', class_='url')

    # Save the titles and URLs of the blog posts to the worksheet
    for i, (title, url) in enumerate(zip(blog_titles, blog_urls)):
        ws.cell(row=i+1, column=1, value=title.text.strip())
        ws.cell(row=i+1, column=2, value=url['href'])

# Save the workbook to an Excel file in the c:\work folder
wb.save(r'c:\work\blog_posts.xlsx')